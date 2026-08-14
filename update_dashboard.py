#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
王敏-达人合作进度看板 · 更新脚本
运行方式：python update_dashboard.py
功能：读取Excel数据，自动生成可分享的HTML看板
"""

import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

EXCEL_FILE = "王敏-达人合作进度看板.xlsx"
HTML_FILE = "index.html"

def read_excel_data():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # --- 数据总览 ---
    ws = wb['数据总览']
    weekly_data = []
    platform_data = []
    kpi_summary = {}

    # 按周统计 (rows 4-6)
    for r in range(4, 7):
        row = [ws.cell(r, c).value for c in range(1, 10)]
        if row[0]:
            weekly_data.append({
                "week": row[0] or "",
                "reach": row[1] or 0,
                "reply": row[2] or 0,
                "connect": row[3] or 0,
                "coop": row[4] or 0,
                "reply_rate": round(row[5] * 100, 1) if row[5] else 0,
                "connect_rate": round(row[6] * 100, 1) if row[6] else 0,
                "coop_rate": round(row[7] * 100, 1) if row[7] else 0,
                "note": row[8] or ""
            })

    # 合计行 (row 7)
    total_row = [ws.cell(7, c).value for c in range(1, 10)]
    weekly_total = {
        "reach": total_row[1] or 0,
        "reply": total_row[2] or 0,
        "connect": total_row[3] or 0,
        "coop": total_row[4] or 0,
        "reply_rate": round(total_row[5] * 100, 1) if total_row[5] else 0,
        "connect_rate": round(total_row[6] * 100, 1) if total_row[6] else 0,
        "coop_rate": round(total_row[7] * 100, 1) if total_row[7] else 0
    }

    # 按平台统计 (rows 13-15)
    for r in range(13, 16):
        row = [ws.cell(r, c).value for c in range(1, 10)]
        if row[0] and (row[1] or 0) > 0:
            platform_data.append({
                "platform": row[0] or "",
                "reach": row[1] or 0,
                "reply": row[2] or 0,
                "connect": row[3] or 0,
                "coop": row[4] or 0,
                "reply_rate": round(row[5] * 100, 1) if row[5] else 0,
                "connect_rate": round(row[6] * 100, 1) if row[6] else 0,
                "coop_rate": round(row[7] * 100, 1) if row[7] else 0,
                "avg_price": row[8] or 0
            })

    # 累计关键指标 (rows 21-27)
    kpi_summary = {
        "total_reach": ws.cell(21, 2).value or 0,
        "total_reply": ws.cell(22, 2).value or 0,
        "total_connect": ws.cell(23, 2).value or 0,
        "total_coop": ws.cell(24, 2).value or 0,
        "reply_rate": round((ws.cell(25, 2).value or 0) * 100, 1),
        "connect_rate": round((ws.cell(26, 2).value or 0) * 100, 1),
        "coop_rate": round((ws.cell(27, 2).value or 0) * 100, 1)
    }

    # --- 建联达人信息表 ---
    ws2 = wb['建联达人信息表']
    influencers = []
    for r in range(2, ws2.max_row + 1):
        name = ws2.cell(r, 1).value
        if not name:
            continue
        followers = ws2.cell(r, 4).value or 0
        # Parse followers to int
        try:
            followers = int(followers)
        except (ValueError, TypeError):
            followers = 0

        influencers.append({
            "name": str(name),
            "platform": ws2.cell(r, 2).value or "",
            "url": ws2.cell(r, 3).value or "",
            "followers": followers,
            "price": str(ws2.cell(r, 5).value or ""),
            "status": ws2.cell(r, 6).value or "",
            "tier": ws2.cell(r, 7).value or "",
            "category": ws2.cell(r, 8).value or "",
            "contact_date": str(ws2.cell(r, 9).value or "")[:10],
            "current_action": ws2.cell(r, 10).value or ""
        })

    # --- 已合作人数：自动从明细表统计（状态=已合作 或 已确认合作），不再依赖手动填写的 B24 ---
    coop_statuses = ("已合作", "已确认合作")
    auto_coop = sum(1 for inf in influencers if inf["status"] in coop_statuses)
    kpi_summary["total_coop"] = auto_coop
    reach = kpi_summary.get("total_reach") or 0
    kpi_summary["coop_rate"] = round(auto_coop / reach * 100, 1) if reach else 0.0

    return {
        "weekly_data": weekly_data,
        "weekly_total": weekly_total,
        "platform_data": platform_data,
        "kpi_summary": kpi_summary,
        "influencers": influencers,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M")
    }


def build_html(data):
    data_json = json.dumps(data, ensure_ascii=False)

    return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>王敏 - 达人合作进度看板</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
:root {{
  --primary: #4f46e5;
  --primary-light: #818cf8;
  --primary-dark: #3730a3;
  --success: #10b981;
  --warning: #f59e0b;
  --danger: #ef4444;
  --info: #3b82f6;
  --gray-50: #f9fafb;
  --gray-100: #f3f4f6;
  --gray-200: #e5e7eb;
  --gray-300: #d1d5db;
  --gray-400: #9ca3af;
  --gray-500: #6b7280;
  --gray-600: #4b5563;
  --gray-700: #374151;
  --gray-800: #1f2937;
  --gray-900: #111827;
}}
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'PingFang SC', 'Hiragino Sans GB', 'Microsoft YaHei', sans-serif;
  background: var(--gray-50);
  color: var(--gray-800);
  line-height: 1.6;
}}
.container {{ max-width: 1400px; margin: 0 auto; padding: 20px; }}

/* Header */
.header {{
  background: linear-gradient(135deg, var(--primary) 0%, var(--primary-dark) 100%);
  color: white;
  padding: 28px 32px;
  border-radius: 16px;
  margin-bottom: 24px;
  display: flex;
  justify-content: space-between;
  align-items: center;
  flex-wrap: wrap;
  gap: 12px;
}}
.header h1 {{ font-size: 24px; font-weight: 700; letter-spacing: 0.5px; }}
.header .subtitle {{ font-size: 14px; opacity: 0.85; margin-top: 4px; }}
.header .update-time {{ font-size: 13px; opacity: 0.75; }}
.header .badge {{
  background: rgba(255,255,255,0.2);
  padding: 6px 14px;
  border-radius: 20px;
  font-size: 13px;
}}

/* KPI Cards */
.kpi-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
  gap: 16px;
  margin-bottom: 24px;
}}
.kpi-card {{
  background: white;
  border-radius: 12px;
  padding: 20px;
  box-shadow: 0 1px 3px rgba(0,0,0,0.08);
  border: 1px solid var(--gray-200);
  transition: transform 0.15s, box-shadow 0.15s;
}}
.kpi-card:hover {{ transform: translateY(-2px); box-shadow: 0 4px 12px rgba(0,0,0,0.1); }}
.kpi-card .label {{ font-size: 13px; color: var(--gray-500); margin-bottom: 6px; }}
.kpi-card .value {{ font-size: 28px; font-weight: 700; line-height: 1.2; }}
.kpi-card .sub {{ font-size: 12px; color: var(--gray-400); margin-top: 4px; }}
.kpi-card.primary .value {{ color: var(--primary); }}
.kpi-card.success .value {{ color: var(--success); }}
.kpi-card.warning .value {{ color: var(--warning); }}
.kpi-card.danger .value {{ color: var(--danger); }}
.kpi-card.info .value {{ color: var(--info); }}

/* Section */
.section {{
  background: white;
  border-radius: 12px;
  padding: 24px;
  margin-bottom: 24px;
  box-shadow: 0 1px 3px rgba(0,0,0,0.08);
  border: 1px solid var(--gray-200);
}}
.section-title {{
  font-size: 18px;
  font-weight: 600;
  margin-bottom: 20px;
  color: var(--gray-800);
  display: flex;
  align-items: center;
  gap: 8px;
}}
.section-title .icon {{ font-size: 20px; }}

/* Chart grid */
.chart-grid {{
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 20px;
}}
@media (max-width: 900px) {{
  .chart-grid {{ grid-template-columns: 1fr; }}
}}
.chart-box {{
  background: var(--gray-50);
  border-radius: 10px;
  padding: 16px;
}}
.chart-box h3 {{
  font-size: 14px;
  font-weight: 600;
  color: var(--gray-600);
  margin-bottom: 12px;
  text-align: center;
}}
.chart-box canvas {{ max-height: 300px; }}

/* Table */
.table-wrap {{
  overflow-x: auto;
  -webkit-overflow-scrolling: touch;
}}
table {{
  width: 100%;
  border-collapse: collapse;
  font-size: 14px;
}}
th {{
  background: var(--gray-100);
  color: var(--gray-600);
  font-weight: 600;
  padding: 10px 12px;
  text-align: left;
  white-space: nowrap;
  position: sticky;
  top: 0;
}}
td {{
  padding: 10px 12px;
  border-bottom: 1px solid var(--gray-200);
  vertical-align: middle;
}}
tr:hover td {{ background: var(--gray-50); }}
.followers-num {{ font-weight: 600; }}

/* Status badges */
.status-badge {{
  display: inline-block;
  padding: 3px 10px;
  border-radius: 12px;
  font-size: 12px;
  font-weight: 500;
  white-space: nowrap;
}}
.status-沟通中 {{ background: #dbeafe; color: #1d4ed8; }}
.status-报价协商中 {{ background: #fef3c7; color: #b45309; }}
.status-已拒绝 {{ background: #fee2e2; color: #dc2626; }}
.status-已确认合作 {{ background: #d1fae5; color: #059669; }}
.status-已合作 {{ background: #d1fae5; color: #059669; }}
.status-新建联 {{ background: #e0e7ff; color: #4338ca; }}
.status-报价中 {{ background: #fce7f3; color: #be185d; }}
.status-已失效 {{ background: #f3f4f6; color: #6b7280; }}

.tier-badge {{
  display: inline-block;
  padding: 2px 8px;
  border-radius: 8px;
  font-size: 12px;
  background: var(--gray-100);
  color: var(--gray-600);
}}
.tier-头部KOL {{ background: #fef3c7; color: #92400e; }}
.tier-腰部KOL {{ background: #dbeafe; color: #1e40af; }}
.tier-尾部KOL {{ background: #e0e7ff; color: #3730a3; }}
.tier-素人KOC {{ background: #f3f4f6; color: #4b5563; }}

/* Funnel */
.funnel {{
  display: flex;
  flex-direction: column;
  align-items: center;
  gap: 4px;
  padding: 10px 0;
}}
.funnel-step {{
  display: flex;
  align-items: center;
  justify-content: center;
  gap: 16px;
  width: 100%;
  max-width: 500px;
  padding: 14px 20px;
  border-radius: 10px;
  color: white;
  font-weight: 600;
  font-size: 15px;
  position: relative;
  transition: transform 0.2s;
}}
.funnel-step:hover {{ transform: scale(1.02); }}
.funnel-step .count {{ font-size: 22px; font-weight: 700; }}
.funnel-step .rate {{ font-size: 13px; opacity: 0.85; }}
.funnel-step.step-1 {{ background: var(--primary); width: 100%; }}
.funnel-step.step-2 {{ background: var(--info); width: 85%; }}
.funnel-step.step-3 {{ background: var(--warning); width: 70%; }}
.funnel-step.step-4 {{ background: var(--success); width: 55%; }}
.funnel-step.step-5 {{ background: var(--danger); width: 40%; }}

/* Footer */
.footer {{
  text-align: center;
  padding: 20px;
  color: var(--gray-400);
  font-size: 13px;
}}
.footer a {{ color: var(--primary); text-decoration: none; }}
</style>
</head>
<body>
<div class="container">

<!-- Header -->
<div class="header">
  <div>
    <h1>📊 达人合作进度看板</h1>
    <div class="subtitle">王敏 · 达人/KOL 外联管理</div>
  </div>
  <div style="text-align:right">
    <div class="update-time">🕐 更新于：{data['updated_at']}</div>
    <div class="badge">数据来源：Excel 自动同步</div>
  </div>
</div>

<!-- KPI Cards -->
<div class="kpi-grid">
  <div class="kpi-card primary">
    <div class="label">📬 累计触达</div>
    <div class="value">{data['kpi_summary']['total_reach']}</div>
    <div class="sub">达人</div>
  </div>
  <div class="kpi-card info">
    <div class="label">💬 累计回复</div>
    <div class="value">{data['kpi_summary']['total_reply']}</div>
    <div class="sub">回复率 {data['kpi_summary']['reply_rate']}%</div>
  </div>
  <div class="kpi-card warning">
    <div class="label">🤝 累计建联</div>
    <div class="value">{data['kpi_summary']['total_connect']}</div>
    <div class="sub">建联率 {data['kpi_summary']['connect_rate']}%</div>
  </div>
  <div class="kpi-card success">
    <div class="label">✅ 合作确认</div>
    <div class="value">{data['kpi_summary']['total_coop']}</div>
    <div class="sub">合作率 {data['kpi_summary']['coop_rate']}%</div>
  </div>
  <div class="kpi-card" style="border-left:3px solid var(--primary);">
    <div class="label">👤 总达人数量</div>
    <div class="value" style="color:var(--primary)">{len(data['influencers'])}</div>
    <div class="sub">人</div>
  </div>
  <div class="kpi-card" style="border-left:3px solid var(--warning);">
    <div class="label">🔄 跟进中</div>
    <div class="value" style="color:var(--warning)">{sum(1 for i in data['influencers'] if i['status'] in ('沟通中','报价协商中','报价中','新建联'))}</div>
    <div class="sub">待推进</div>
  </div>
</div>

<!-- Charts Row 1: Weekly Trend + Funnel -->
<div class="section">
  <div class="section-title"><span class="icon">📈</span> 合作进度总览</div>
  <div class="chart-grid">
    <div class="chart-box">
      <h3>按周趋势（触达/回复/建联）</h3>
      <canvas id="weeklyChart"></canvas>
    </div>
    <div class="chart-box">
      <h3>合作状态分布</h3>
      <canvas id="statusChart"></canvas>
    </div>
  </div>
</div>

<!-- Charts Row 2: Tier + Platform -->
<div class="section">
  <div class="section-title"><span class="icon">📊</span> 达人结构分析</div>
  <div class="chart-grid">
    <div class="chart-box">
      <h3>达人层级分布</h3>
      <canvas id="tierChart"></canvas>
    </div>
    <div class="chart-box">
      <h3>各平台效能对比</h3>
      <canvas id="platformChart"></canvas>
    </div>
  </div>
</div>

<!-- Influencer Table -->
<div class="section">
  <div class="section-title"><span class="icon">📋</span> 达人明细表 <span style="font-size:13px;font-weight:400;color:var(--gray-400);margin-left:8px;">共 {len(data['influencers'])} 人</span></div>
  <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th>达人名称</th>
          <th>平台</th>
          <th>粉丝量</th>
          <th>层级</th>
          <th>内容类别</th>
          <th>合作状态</th>
          <th>触达日期</th>
          <th>当前状态</th>
        </tr>
      </thead>
      <tbody>
{''.join(f'''        <tr>
          <td><strong>{inf['name']}</strong></td>
          <td>{inf['platform']}</td>
          <td class="followers-num">{format_followers(inf['followers'])}</td>
          <td><span class="tier-badge tier-{inf['tier']}">{inf['tier']}</span></td>
          <td>{inf['category']}</td>
          <td><span class="status-badge status-{inf['status']}">{inf['status']}</span></td>
          <td>{inf['contact_date']}</td>
          <td style="color:var(--gray-500);font-size:13px;">{inf['current_action']}</td>
        </tr>''' for inf in data['influencers'])}
      </tbody>
    </table>
  </div>
</div>

<!-- Footer -->
<div class="footer">
  <p>📊 王敏 · 达人合作进度看板 &nbsp;|&nbsp; 数据自动同步自 Excel &nbsp;|&nbsp; 更新于 {data['updated_at']}</p>
  <p style="margin-top:6px;">💡 更新 Excel 数据后，运行 <code>python update_dashboard.py</code> 即可刷新本看板</p>
</div>

</div>

<script>
const DATA = {data_json};

// --- KPI 统计 ---
const statuses = {{}};
const tiers = {{}};
DATA.influencers.forEach(inf => {{
  statuses[inf.status] = (statuses[inf.status] || 0) + 1;
  tiers[inf.tier] = (tiers[inf.tier] || 0) + 1;
}});

const STATUS_COLORS = {{
  '新建联': '#4338ca',
  '沟通中': '#1d4ed8',
  '报价中': '#be185d',
  '报价协商中': '#b45309',
  '已确认合作': '#059669',
  '已合作': '#059669',
  '已拒绝': '#dc2626',
  '已失效': '#6b7280'
}};
const STATUS_ORDER = ['新建联','沟通中','报价中','报价协商中','已合作','已确认合作','已拒绝','已失效'];

// --- 1. Weekly Chart ---
const weeks = DATA.weekly_data.map(d => d.week.replace(/第|周.*/g,'').trim());
new Chart(document.getElementById('weeklyChart'), {{
  type: 'bar',
  data: {{
    labels: DATA.weekly_data.map(d => d.week),
    datasets: [
      {{
        label: '触达人数',
        data: DATA.weekly_data.map(d => d.reach),
        backgroundColor: 'rgba(79,70,229,0.7)',
        borderColor: '#4f46e5',
        borderWidth: 1
      }},
      {{
        label: '回复人数',
        data: DATA.weekly_data.map(d => d.reply),
        backgroundColor: 'rgba(59,130,246,0.7)',
        borderColor: '#3b82f6',
        borderWidth: 1
      }},
      {{
        label: '建联人数',
        data: DATA.weekly_data.map(d => d.connect),
        backgroundColor: 'rgba(245,158,11,0.7)',
        borderColor: '#f59e0b',
        borderWidth: 1
      }}
    ]
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: true,
    plugins: {{ legend: {{ position: 'bottom', labels: {{ boxWidth: 12, padding: 12 }} }} }},
    scales: {{
      y: {{ beginAtZero: true, grid: {{ color: 'rgba(0,0,0,0.05)' }} }},
      x: {{ grid: {{ display: false }} }}
    }}
  }}
}});

// --- 2. Status Chart ---
const sortedStatuses = STATUS_ORDER.filter(s => statuses[s]).map(s => ({{ status: s, count: statuses[s] }}));
new Chart(document.getElementById('statusChart'), {{
  type: 'doughnut',
  data: {{
    labels: sortedStatuses.map(s => s.status + ' (' + s.count + ')'),
    datasets: [{{
      data: sortedStatuses.map(s => s.count),
      backgroundColor: sortedStatuses.map(s => STATUS_COLORS[s.status] || '#9ca3af'),
      borderWidth: 2,
      borderColor: '#fff'
    }}]
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: true,
    plugins: {{
      legend: {{ position: 'bottom', labels: {{ boxWidth: 12, padding: 10, font: {{ size: 11 }} }} }}
    }}
  }}
}});

// --- 3. Tier Chart ---
const tierOrder = ['头部KOL','腰部KOL','尾部KOL','素人KOC'];
const sortedTiers = tierOrder.filter(t => tiers[t]).map(t => ({{ tier: t, count: tiers[t] }}));
const TIER_COLORS = ['#f59e0b','#3b82f6','#818cf8','#9ca3af'];
new Chart(document.getElementById('tierChart'), {{
  type: 'pie',
  data: {{
    labels: sortedTiers.map(t => t.tier + ' (' + t.count + ')'),
    datasets: [{{
      data: sortedTiers.map(t => t.count),
      backgroundColor: TIER_COLORS.slice(0, sortedTiers.length),
      borderWidth: 2,
      borderColor: '#fff'
    }}]
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: true,
    plugins: {{
      legend: {{ position: 'bottom', labels: {{ boxWidth: 12, padding: 10, font: {{ size: 11 }} }} }}
    }}
  }}
}});

// --- 4. Platform Chart ---
const platforms = DATA.platform_data.length ? DATA.platform_data : [{{platform:'TikTok',reach:0,reply:0,connect:0,coop:0,reply_rate:0,connect_rate:0,coop_rate:0,avg_price:0}}];
new Chart(document.getElementById('platformChart'), {{
  type: 'radar',
  data: {{
    labels: ['触达','回复','建联','合作'],
    datasets: platforms.map((p, i) => ({{
      label: p.platform,
      data: [p.reach, p.reply, p.connect, p.coop],
      backgroundColor: ['rgba(79,70,229,0.1)','rgba(16,185,129,0.1)','rgba(245,158,11,0.1)'][i] || 'rgba(59,130,246,0.1)',
      borderColor: ['#4f46e5','#10b981','#f59e0b','#3b82f6'][i] || '#3b82f6',
      borderWidth: 2,
      pointBackgroundColor: ['#4f46e5','#10b981','#f59e0b','#3b82f6'][i] || '#3b82f6'
    }}))
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: true,
    plugins: {{ legend: {{ position: 'bottom', labels: {{ boxWidth: 12, padding: 10 }} }} }},
    scales: {{ r: {{ beginAtZero: true, grid: {{ color: 'rgba(0,0,0,0.06)' }} }} }}
  }}
}});
</script>
</body>
</html>'''

def format_followers(n):
    if n >= 1000000:
        return f"{n/1000000:.1f}M"
    elif n >= 1000:
        return f"{n/1000:.0f}K"
    return str(n)

def main():
    import sys; sys.stdout.reconfigure(encoding='utf-8')
    print(f"📖 读取数据：{EXCEL_FILE}")
    data = read_excel_data()
    print(f"   ✅ 达人数据：{len(data['influencers'])} 人")
    print(f"   ✅ 周次数据：{len(data['weekly_data'])} 周")
    print(f"   ✅ 平台数据：{len(data['platform_data'])} 个平台")

    html = build_html(data)
    with open(HTML_FILE, 'w', encoding='utf-8') as f:
        f.write(html)

    full_path = os.path.abspath(HTML_FILE)
    print(f"\n🎉 看板已生成：{full_path}")
    print(f"📎 文件大小：{os.path.getsize(HTML_FILE)/1024:.0f} KB")
    print(f"\n💡 下次更新 Excel 后，运行以下命令即可刷新看板：")
    print(f"   python update_dashboard.py")

if __name__ == '__main__':
    main()